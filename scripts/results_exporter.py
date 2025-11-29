#!/usr/bin/env python
"""
Results Exporter

Export orchestrator results to multiple formats (JSON, CSV, Excel, Markdown, HTML).

Features:
- Multiple output formats
- Aggregation and filtering
- Time-series data export
- Report generation
- Customizable templates

Usage examples (PowerShell):
  python .\\scripts\\results_exporter.py --source autotrain --format csv
  python .\\scripts\\results_exporter.py --source quantum_autorun --format markdown
  python .\\scripts\\results_exporter.py --all --format excel
  python .\\scripts\\results_exporter.py --compare autotrain quantum_autorun --format html
"""

from __future__ import annotations

import argparse
import csv
import json
import sys
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional

REPO_ROOT = Path(__file__).resolve().parents[1]
DATA_OUT = REPO_ROOT / "data_out"
EXPORT_DIR = REPO_ROOT / "exports"


@dataclass
class ExportConfig:
    """Configuration for result export."""
    source: str
    format: str
    output_file: Optional[Path] = None
    include_metadata: bool = True
    filter_status: Optional[str] = None


class ResultsExporter:
    """Exports orchestrator results to various formats."""
    
    def __init__(self):
        self.export_dir = EXPORT_DIR
        self.export_dir.mkdir(parents=True, exist_ok=True)
    
    def load_results(self, orchestrator: str) -> Optional[Dict]:
        """Load results from orchestrator status.json."""
        status_file = DATA_OUT / orchestrator / "status.json"
        
        if not status_file.exists():
            print(f"[exporter] Status file not found: {status_file}")
            return None
        
        with status_file.open("r") as f:
            return json.load(f)
    
    def export_json(self, data: Dict, output_file: Path):
        """Export to JSON format."""
        with output_file.open("w") as f:
            json.dump(data, f, indent=2)
        print(f"[exporter] Exported JSON to: {output_file}")
    
    def export_csv(self, data: Dict, output_file: Path):
        """Export to CSV format."""
        jobs = data.get("jobs", [])
        
        if not jobs:
            print("[exporter] No jobs to export")
            return
        
        with output_file.open("w", newline="") as f:
            # Determine all keys from jobs
            fieldnames = set()
            for job in jobs:
                fieldnames.update(job.keys())
            
            fieldnames = sorted(fieldnames)
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            
            writer.writeheader()
            for job in jobs:
                # Flatten nested dicts
                row = {}
                for key, value in job.items():
                    if isinstance(value, dict):
                        row[key] = json.dumps(value)
                    elif isinstance(value, list):
                        row[key] = ", ".join(str(v) for v in value)
                    else:
                        row[key] = value
                writer.writerow(row)
        
        print(f"[exporter] Exported CSV to: {output_file}")
    
    def export_markdown(self, data: Dict, output_file: Path, orchestrator: str):
        """Export to Markdown format."""
        with output_file.open("w") as f:
            # Header
            f.write(f"# {orchestrator.title()} Results\n\n")
            
            if "generated_at" in data:
                f.write(f"**Generated:** {data['generated_at']}\n\n")
            
            # Summary
            f.write("## Summary\n\n")
            if "total_jobs" in data:
                f.write(f"- **Total Jobs:** {data['total_jobs']}\n")
            if "succeeded" in data:
                f.write(f"- **Succeeded:** {data['succeeded']}\n")
            if "failed" in data:
                f.write(f"- **Failed:** {data['failed']}\n")
            if "total_duration" in data:
                f.write(f"- **Total Duration:** {data['total_duration']:.1f}s\n")
            f.write("\n")
            
            # Jobs table
            jobs = data.get("jobs", [])
            if jobs:
                f.write("## Job Results\n\n")
                f.write("| Name | Status | Duration | Completed |\n")
                f.write("|------|--------|----------|----------|\n")
                
                for job in jobs:
                    status_icon = "✓" if job["status"] == "succeeded" else "✗"
                    name = job.get("name", "unknown")
                    status = job.get("status", "unknown")
                    duration = job.get("duration", 0)
                    completed = job.get("completed_at", "-")
                    
                    f.write(f"| {name} | {status_icon} {status} | {duration:.1f}s | {completed} |\n")
                
                f.write("\n")
            
            # Aggregated metrics
            if "aggregated" in data:
                f.write("## Aggregated Metrics\n\n")
                f.write("```json\n")
                f.write(json.dumps(data["aggregated"], indent=2))
                f.write("\n```\n\n")
        
        print(f"[exporter] Exported Markdown to: {output_file}")
    
    def export_html(self, data: Dict, output_file: Path, orchestrator: str):
        """Export to HTML format."""
        with output_file.open("w") as f:
            f.write("<!DOCTYPE html>\n")
            f.write("<html>\n<head>\n")
            f.write(f"<title>{orchestrator.title()} Results</title>\n")
            f.write("<style>\n")
            f.write("body { font-family: Arial, sans-serif; margin: 20px; }\n")
            f.write("table { border-collapse: collapse; width: 100%; margin: 20px 0; }\n")
            f.write("th, td { border: 1px solid #ddd; padding: 8px; text-align: left; }\n")
            f.write("th { background-color: #4CAF50; color: white; }\n")
            f.write(".success { color: green; }\n")
            f.write(".failure { color: red; }\n")
            f.write(".summary { background-color: #f0f0f0; padding: 10px; margin: 10px 0; }\n")
            f.write("</style>\n")
            f.write("</head>\n<body>\n")
            
            # Title
            f.write(f"<h1>{orchestrator.title()} Results</h1>\n")
            
            if "generated_at" in data:
                f.write(f"<p><strong>Generated:</strong> {data['generated_at']}</p>\n")
            
            # Summary
            f.write("<div class='summary'>\n<h2>Summary</h2>\n<ul>\n")
            if "total_jobs" in data:
                f.write(f"<li><strong>Total Jobs:</strong> {data['total_jobs']}</li>\n")
            if "succeeded" in data:
                f.write(f"<li><strong>Succeeded:</strong> {data['succeeded']}</li>\n")
            if "failed" in data:
                f.write(f"<li><strong>Failed:</strong> {data['failed']}</li>\n")
            if "total_duration" in data:
                f.write(f"<li><strong>Total Duration:</strong> {data['total_duration']:.1f}s</li>\n")
            f.write("</ul>\n</div>\n")
            
            # Jobs table
            jobs = data.get("jobs", [])
            if jobs:
                f.write("<h2>Job Results</h2>\n")
                f.write("<table>\n<thead>\n<tr>\n")
                f.write("<th>Name</th><th>Status</th><th>Duration</th><th>Completed</th>\n")
                f.write("</tr>\n</thead>\n<tbody>\n")
                
                for job in jobs:
                    status = job.get("status", "unknown")
                    status_class = "success" if status == "succeeded" else "failure"
                    status_icon = "✓" if status == "succeeded" else "✗"
                    
                    f.write("<tr>\n")
                    f.write(f"<td>{job.get('name', 'unknown')}</td>\n")
                    f.write(f"<td class='{status_class}'>{status_icon} {status}</td>\n")
                    f.write(f"<td>{job.get('duration', 0):.1f}s</td>\n")
                    f.write(f"<td>{job.get('completed_at', '-')}</td>\n")
                    f.write("</tr>\n")
                
                f.write("</tbody>\n</table>\n")
            
            f.write("</body>\n</html>\n")
        
        print(f"[exporter] Exported HTML to: {output_file}")
    
    def export_excel(self, data: Dict, output_file: Path):
        """Export to Excel format (requires openpyxl)."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            print("[exporter] ERROR: openpyxl not installed. Install with: pip install openpyxl")
            return
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Results"
        
        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
        
        jobs = data.get("jobs", [])
        if not jobs:
            print("[exporter] No jobs to export")
            return
        
        # Determine headers
        headers = sorted(set().union(*[job.keys() for job in jobs]))
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        # Write data
        for row, job in enumerate(jobs, 2):
            for col, header in enumerate(headers, 1):
                value = job.get(header, "")
                if isinstance(value, (dict, list)):
                    value = json.dumps(value)
                ws.cell(row=row, column=col, value=value)
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        wb.save(output_file)
        print(f"[exporter] Exported Excel to: {output_file}")
    
    def export_comparison(self, orchestrators: List[str], format: str, output_file: Path):
        """Export comparison of multiple orchestrators."""
        comparison = {
            "generated_at": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
            "orchestrators": []
        }
        
        for orch in orchestrators:
            data = self.load_results(orch)
            if data:
                summary = {
                    "name": orch,
                    "total_jobs": data.get("total_jobs", 0),
                    "succeeded": data.get("succeeded", 0),
                    "failed": data.get("failed", 0),
                    "total_duration": data.get("total_duration", 0),
                }
                comparison["orchestrators"].append(summary)
        
        if format == "json":
            self.export_json(comparison, output_file)
        elif format == "markdown":
            self._export_comparison_markdown(comparison, output_file)
        elif format == "html":
            self._export_comparison_html(comparison, output_file)
        elif format == "csv":
            with output_file.open("w", newline="") as f:
                writer = csv.DictWriter(f, fieldnames=["name", "total_jobs", "succeeded", "failed", "total_duration"])
                writer.writeheader()
                writer.writerows(comparison["orchestrators"])
            print(f"[exporter] Exported comparison CSV to: {output_file}")
    
    def _export_comparison_markdown(self, comparison: Dict, output_file: Path):
        """Export orchestrator comparison as Markdown."""
        with output_file.open("w") as f:
            f.write("# Orchestrator Comparison\n\n")
            f.write(f"**Generated:** {comparison['generated_at']}\n\n")
            
            f.write("| Orchestrator | Total Jobs | Succeeded | Failed | Duration |\n")
            f.write("|--------------|------------|-----------|--------|----------|\n")
            
            for orch in comparison["orchestrators"]:
                f.write(f"| {orch['name']} | {orch['total_jobs']} | {orch['succeeded']} | {orch['failed']} | {orch['total_duration']:.1f}s |\n")
        
        print(f"[exporter] Exported comparison Markdown to: {output_file}")
    
    def _export_comparison_html(self, comparison: Dict, output_file: Path):
        """Export orchestrator comparison as HTML."""
        with output_file.open("w") as f:
            f.write("<!DOCTYPE html>\n<html>\n<head>\n")
            f.write("<title>Orchestrator Comparison</title>\n")
            f.write("<style>\n")
            f.write("body { font-family: Arial, sans-serif; margin: 20px; }\n")
            f.write("table { border-collapse: collapse; width: 100%; }\n")
            f.write("th, td { border: 1px solid #ddd; padding: 8px; text-align: left; }\n")
            f.write("th { background-color: #4CAF50; color: white; }\n")
            f.write("</style>\n</head>\n<body>\n")
            
            f.write("<h1>Orchestrator Comparison</h1>\n")
            f.write(f"<p><strong>Generated:</strong> {comparison['generated_at']}</p>\n")
            
            f.write("<table>\n<thead>\n<tr>\n")
            f.write("<th>Orchestrator</th><th>Total Jobs</th><th>Succeeded</th><th>Failed</th><th>Duration</th>\n")
            f.write("</tr>\n</thead>\n<tbody>\n")
            
            for orch in comparison["orchestrators"]:
                f.write("<tr>\n")
                f.write(f"<td>{orch['name']}</td>\n")
                f.write(f"<td>{orch['total_jobs']}</td>\n")
                f.write(f"<td>{orch['succeeded']}</td>\n")
                f.write(f"<td>{orch['failed']}</td>\n")
                f.write(f"<td>{orch['total_duration']:.1f}s</td>\n")
                f.write("</tr>\n")
            
            f.write("</tbody>\n</table>\n</body>\n</html>\n")
        
        print(f"[exporter] Exported comparison HTML to: {output_file}")


def main():
    ap = argparse.ArgumentParser(description="Results Exporter")
    ap.add_argument("--source", help="Source orchestrator (autotrain, quantum_autorun, evaluation_autorun)")
    ap.add_argument("--all", action="store_true", help="Export all orchestrators")
    ap.add_argument("--compare", nargs="+", help="Compare multiple orchestrators")
    ap.add_argument("--format", required=True, choices=["json", "csv", "excel", "markdown", "html"], help="Export format")
    ap.add_argument("--output", type=Path, help="Output file")
    ap.add_argument("--filter-status", choices=["succeeded", "failed"], help="Filter jobs by status")
    args = ap.parse_args()
    
    exporter = ResultsExporter()
    
    # Generate default output filename
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    if not args.output:
        if args.compare:
            basename = f"comparison_{timestamp}"
        elif args.all:
            basename = f"all_orchestrators_{timestamp}"
        elif args.source:
            basename = f"{args.source}_{timestamp}"
        else:
            basename = f"export_{timestamp}"
        
        extension = {"json": ".json", "csv": ".csv", "excel": ".xlsx", "markdown": ".md", "html": ".html"}
        args.output = exporter.export_dir / f"{basename}{extension[args.format]}"
    
    # Handle comparison
    if args.compare:
        exporter.export_comparison(args.compare, args.format, args.output)
        return
    
    # Handle all orchestrators
    if args.all:
        orchestrators = ["autotrain", "quantum_autorun", "evaluation_autorun"]
        exporter.export_comparison(orchestrators, args.format, args.output)
        return
    
    # Single orchestrator export
    if args.source:
        data = exporter.load_results(args.source)
        if not data:
            sys.exit(1)
        
        # Apply filter
        if args.filter_status and "jobs" in data:
            data["jobs"] = [j for j in data["jobs"] if j.get("status") == args.filter_status]
        
        # Export based on format
        if args.format == "json":
            exporter.export_json(data, args.output)
        elif args.format == "csv":
            exporter.export_csv(data, args.output)
        elif args.format == "excel":
            exporter.export_excel(data, args.output)
        elif args.format == "markdown":
            exporter.export_markdown(data, args.output, args.source)
        elif args.format == "html":
            exporter.export_html(data, args.output, args.source)
        
        return
    
    # Default: show help
    ap.print_help()


if __name__ == "__main__":
    main()
